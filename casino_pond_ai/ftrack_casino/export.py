"""
Multi-format export for FastTrack issues

Supports: Excel, CSV, JIRA CSV, HTML
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any


class IssueExporter:
    """Export issues to various formats"""

    def __init__(self, database):
        self.db = database

    def export_to_csv(self, issues: List[Dict[str, Any]], filename: str):
        """Export issues to CSV"""
        if not issues:
            return

        # Define columns
        columns = ['ID', 'Title', 'Status', 'Severity', 'Stage', 'Assignee',
                   'Assigner', 'Created', 'Due Date', 'Blocks', 'Run Directory', 'Description']

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()

            for issue in issues:
                # Parse modules from JSON
                modules_str = ', '.join(json.loads(issue.get('modules', '[]')))

                writer.writerow({
                    'ID': issue.get('id', ''),
                    'Title': issue.get('title', ''),
                    'Status': issue.get('status', ''),
                    'Severity': issue.get('severity', ''),
                    'Stage': issue.get('stage', ''),
                    'Assignee': issue.get('assignee', ''),
                    'Assigner': issue.get('assigner', ''),
                    'Created': issue.get('created_at', ''),
                    'Due Date': issue.get('due_date', ''),
                    'Blocks': modules_str,
                    'Run Directory': issue.get('run_id', ''),
                    'Description': issue.get('description', '')
                })

    def export_to_excel(self, issues: List[Dict[str, Any]], filename: str):
        """Export to Excel with formatting"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            print("openpyxl not installed. Falling back to CSV export.")
            self.export_to_csv(issues, filename.replace('.xlsx', '.csv'))
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Issues"

        # Headers
        headers = ['ID', 'Title', 'Status', 'Severity', 'Stage', 'Assignee',
                   'Assigner', 'Created', 'Due Date', 'Blocks', 'Run Directory']
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            cell.font = Font(bold=True)

        # Data rows
        for issue in issues:
            modules_str = ', '.join(json.loads(issue.get('modules', '[]')))

            row = [
                issue.get('id', ''),
                issue.get('title', ''),
                issue.get('status', ''),
                issue.get('severity', ''),
                issue.get('stage', ''),
                issue.get('assignee', ''),
                issue.get('assigner', ''),
                issue.get('created_at', ''),
                issue.get('due_date', ''),
                modules_str,
                issue.get('run_id', '')
            ]
            ws.append(row)

            # Color code by severity
            severity = issue.get('severity', '')
            severity_colors = {
                'Critical': 'FFDC3545',
                'Major': 'FFFD7E14',
                'Minor': 'FFFFC107',
                'Enhancement': 'FF28A745',
                'Info': 'FFE9ECEF'
            }

            if severity in severity_colors:
                fill = PatternFill(start_color=severity_colors[severity],
                                 end_color=severity_colors[severity],
                                 fill_type='solid')
                for cell in ws[ws.max_row]:
                    cell.fill = fill

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(filename)

    def export_to_jira_csv(self, issues: List[Dict[str, Any]], filename: str):
        """Export in JIRA-compatible CSV format"""
        columns = ['Summary', 'Issue Type', 'Priority', 'Status', 'Assignee',
                   'Description', 'Labels', 'Created', 'Due Date']

        severity_to_priority = {
            'Critical': 'Highest',
            'Major': 'High',
            'Minor': 'Medium',
            'Enhancement': 'Low',
            'Info': 'Lowest'
        }

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()

            for issue in issues:
                modules_str = ','.join(json.loads(issue.get('modules', '[]')))

                writer.writerow({
                    'Summary': issue.get('title', ''),
                    'Issue Type': 'Task',
                    'Priority': severity_to_priority.get(issue.get('severity', ''), 'Medium'),
                    'Status': issue.get('status', ''),
                    'Assignee': issue.get('assignee', ''),
                    'Description': issue.get('description', ''),
                    'Labels': f"{issue.get('stage', '')},{modules_str}",
                    'Created': issue.get('created_at', ''),
                    'Due Date': issue.get('due_date', '')
                })

    def export_to_html(self, issues: List[Dict[str, Any]], filename: str, project_name: str = ""):
        """Export to HTML with styling"""
        html = f"""<!DOCTYPE html>
<html>
<head>
    <title>Issue Export - {project_name}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #2c3e50; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #E6E6FA; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .status-open {{ background-color: #f8d7da; }}
        .status-in-progress {{ background-color: #fff3cd; }}
        .status-resolved {{ background-color: #d4edda; }}
        .status-closed {{ background-color: #d1ecf1; }}
        .severity-critical {{ background-color: #dc3545; color: white; }}
        .severity-major {{ background-color: #fd7e14; color: white; }}
        .severity-minor {{ background-color: #ffc107; }}
        .severity-enhancement {{ background-color: #28a745; color: white; }}
        .overdue {{ color: red; font-weight: bold; }}
    </style>
</head>
<body>
    <h1>Issue Export - {project_name}</h1>
    <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    <p>Total Issues: {len(issues)}</p>

    <table>
        <tr>
            <th>ID</th>
            <th>Title</th>
            <th>Status</th>
            <th>Severity</th>
            <th>Stage</th>
            <th>Assignee</th>
            <th>Created</th>
            <th>Due Date</th>
        </tr>
"""

        for issue in issues:
            status = issue.get('status', '').lower().replace(' ', '-')
            severity = issue.get('severity', '').lower()
            due_date = issue.get('due_date', '')

            # Check if overdue
            overdue_class = ""
            if due_date:
                try:
                    due_dt = datetime.fromisoformat(due_date.split()[0])
                    if due_dt < datetime.now() and issue.get('status') not in ['Resolved', 'Closed']:
                        overdue_class = "overdue"
                except:
                    pass

            html += f"""
        <tr>
            <td>{issue.get('id', '')}</td>
            <td>{issue.get('title', '')}</td>
            <td class="status-{status}">{issue.get('status', '')}</td>
            <td class="severity-{severity}">{issue.get('severity', '')}</td>
            <td>{issue.get('stage', '')}</td>
            <td>{issue.get('assignee', '')}</td>
            <td>{issue.get('created_at', '')}</td>
            <td class="{overdue_class}">{due_date}</td>
        </tr>
"""

        html += """
    </table>
</body>
</html>"""

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html)
